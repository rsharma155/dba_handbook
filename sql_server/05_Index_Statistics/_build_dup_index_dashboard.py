"""Build Executive_Dashboard + Charts sheets into the duplicate index Excel report.

Charts sheet: each chart has an adjacent editable data table (easy to tweak).
Modern doughnut/bar layouts with short labels to avoid text overlap.
"""
import warnings
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
import pandas as pd
import numpy as np
from datetime import datetime

path = r'D:\Mac_bak\AI_Code\sql_optima\dba_essential_scripts\sql_server\05_Index_Statistics\SQLHelps_DTRG_Duplicate_Index_Analysis.xlsx'

# ── Palette (modern, high-contrast, not purple-default) ───────────
PALETTE = ['1B6CA8', 'E07A3D', '2A9D8F', 'E63946', '264653', 'F4A261', '457B9D', 'BC4749']
PALETTE_ALT = ['1B6CA8', 'E07A3D']

# ── Load source data ──────────────────────────────────────────────
diag = pd.read_excel(path, sheet_name='Collection_diagnosis')
dup = pd.read_excel(path, sheet_name='Duplicate_Index_details')
summary_dbs = pd.read_excel(path, sheet_name='Summary_DBs')
big = pd.read_excel(path, sheet_name='Big Tables')
consol = pd.read_excel(path, sheet_name='Index_Consolidation')
consol_sum = pd.read_excel(path, sheet_name='Consolidation_Summary')

dup['keep_reads_f'] = pd.to_numeric(dup['drop_reads'], errors='coerce').fillna(0)
dup['drop_reads_f'] = pd.to_numeric(dup['keep_writes'], errors='coerce').fillna(0)
dup['keep_usage_f'] = dup['drop_usage_status'].astype(str)
dup['drop_usage_f'] = dup['keep_size_mb'].astype(str)
dup['keep_size_f'] = pd.to_numeric(dup['drop_size_mb'], errors='coerce').fillna(0)
dup['drop_size_f'] = pd.to_numeric(dup['suggested_action'], errors='coerce').fillna(0)
dup['action_f'] = dup['drop_ddl'].astype(str)


def safety_bucket(u):
    u = str(u)
    if 'NO USAGE' in u or 'WRITE ONLY' in u or 'UNUSED' in u:
        return 'Safe (unused)'
    if 'USED' in u:
        return 'Review (has reads)'
    return 'Unknown'


dup['safety'] = dup['drop_usage_f'].apply(safety_bucket)


def env_group(db):
    db = str(db)
    if db.startswith(('_Clone', '_Release', '__')):
        return 'Dev/Clone'
    if db.startswith('CLI_'):
        return 'Clinic'
    if db.startswith('HOST_'):
        return 'Hospital HOST'
    if db.startswith('HOS_'):
        return 'Hospital HOS'
    if db.startswith('LAB_'):
        return 'Lab'
    if db.startswith('RES_'):
        return 'Restaurant'
    if db.startswith('PHA_') or 'Pharma' in db or 'pharma' in db.lower():
        return 'Pharma'
    return 'Other'


TYPE_SHORT = {
    'EXACT_DUPLICATE': 'Exact Dup',
    'DUPLICATE_KEY': 'Dup Key',
    'REDUNDANT_LEFT_PREFIX': 'Left Prefix',
}
CONSOL_SHORT = {
    'LEADING_KEY_REVIEW': 'Leading Key',
    'PREFIX_SUBSUMED': 'Prefix Subsumed',
    'DROP_REDUNDANT': 'Drop Redundant',
}


def short_name(s, n=26):
    s = str(s)
    return s if len(s) <= n else s[: n - 1] + '…'


def short_pattern(table, dtype):
    t = str(table).replace('dtrg_hos_', '')
    return short_name(f'{t} - {TYPE_SHORT.get(dtype, dtype)}', 34)


dup['env'] = dup['database_name'].apply(env_group)
summary_dbs['env'] = summary_dbs['database_name'].apply(env_group)

diag['is_merge'] = diag['status_message'].astype(str).str.startswith('MERGE')
ok = diag[~diag['is_merge']].copy()
ok['indexes_scanned'] = pd.to_numeric(ok['indexes_scanned'], errors='coerce').fillna(0)
ok['pairs_found'] = pd.to_numeric(ok['pairs_found'], errors='coerce').fillna(0)

# ── Aggregations ──────────────────────────────────────────────────
kpi_dbs_scanned = int(ok['database_name'].nunique())
kpi_indexes = int(ok['indexes_scanned'].sum())
kpi_pairs = int(dup.shape[0])
kpi_dbs_with_dups = int(dup['database_name'].nunique())
kpi_reclaim_dup = float(summary_dbs['reclaimable_drop_candidate_mb'].sum())
kpi_safe = int((dup['safety'].str.startswith('Safe')).sum())
kpi_review = int((dup['safety'].str.startswith('Review')).sum())
kpi_consol = int(consol.shape[0])
kpi_consol_reclaim = float(pd.to_numeric(consol['est_space_reclaim_mb'], errors='coerce').sum())
kpi_exact = int((dup['duplicate_type'] == 'EXACT_DUPLICATE').sum())
kpi_dupkey = int((dup['duplicate_type'] == 'DUPLICATE_KEY').sum())
kpi_prefix = int((dup['duplicate_type'] == 'REDUNDANT_LEFT_PREFIX').sum())

by_type = (
    summary_dbs.groupby('duplicate_type', as_index=False)
    .agg(pair_count=('pair_count', 'sum'), reclaim_mb=('reclaimable_drop_candidate_mb', 'sum'))
    .sort_values('pair_count', ascending=False)
)
by_type['label'] = by_type['duplicate_type'].map(TYPE_SHORT).fillna(by_type['duplicate_type'])

by_safety = (
    dup.groupby('safety', as_index=False)
    .agg(pairs=('database_name', 'size'), reclaim_mb=('drop_size_f', 'sum'))
    .sort_values('pairs', ascending=False)
)

by_env = (
    dup.groupby('env', as_index=False)
    .agg(pairs=('database_name', 'size'), databases=('database_name', 'nunique'), reclaim_mb=('drop_size_f', 'sum'))
    .sort_values('pairs', ascending=False)
)

top_dbs = (
    summary_dbs.groupby('database_name', as_index=False)
    .agg(pairs=('pair_count', 'sum'), reclaim_mb=('reclaimable_drop_candidate_mb', 'sum'))
    .sort_values('reclaim_mb', ascending=False)
    .head(12)
)
top_dbs['label'] = top_dbs['database_name'].apply(lambda x: short_name(x, 28))

patterns = (
    dup.groupby(['table_name', 'duplicate_type', 'index_keep', 'index_drop_candidate'], as_index=False)
    .agg(databases_affected=('database_name', 'nunique'), reclaim_mb=('drop_size_f', 'sum'))
    .sort_values('databases_affected', ascending=False)
    .head(8)
)
patterns['label'] = patterns.apply(lambda r: short_pattern(r['table_name'], r['duplicate_type']), axis=1)

priority = dup.copy()
priority['priority_score'] = 0.0
priority.loc[priority['duplicate_type'] == 'EXACT_DUPLICATE', 'priority_score'] += 100
priority.loc[priority['duplicate_type'] == 'REDUNDANT_LEFT_PREFIX', 'priority_score'] += 60
priority.loc[priority['duplicate_type'] == 'DUPLICATE_KEY', 'priority_score'] += 40
priority.loc[priority['safety'].str.startswith('Safe'), 'priority_score'] += 50
priority.loc[priority['safety'].str.startswith('Review'), 'priority_score'] -= 20
priority['priority_score'] += priority['drop_size_f'] * 2
priority = priority.sort_values(['priority_score', 'drop_size_f'], ascending=False)
top_actions = priority.head(20)[
    ['database_name', 'table_name', 'duplicate_type', 'index_drop_candidate', 'index_keep',
     'drop_size_f', 'drop_usage_f', 'safety', 'action_f', 'priority_score']
].copy()

consol_by_type = (
    consol_sum.groupby('suggestion_type', as_index=False)
    .agg(
        suggestions=('suggestion_count', 'sum'),
        actionable=('actionable_count', 'sum'),
        reclaim_mb=('est_space_reclaim_mb', 'sum'),
    )
    .sort_values('reclaim_mb', ascending=False)
)
consol_by_type['label'] = consol_by_type['suggestion_type'].map(CONSOL_SHORT).fillna(
    consol_by_type['suggestion_type']
)

top_consol_dbs = (
    consol_sum.groupby('database_name', as_index=False)
    .agg(suggestions=('suggestion_count', 'sum'), reclaim_mb=('est_space_reclaim_mb', 'sum'))
    .sort_values('reclaim_mb', ascending=False)
    .head(10)
)

big_top = big.nlargest(12, 'table_size_mb')[
    ['database_name', 'schema_name', 'table_name', 'table_rows', 'table_size_mb', 'index_count', 'nc_index_count']
].copy()

# ── Styles ────────────────────────────────────────────────────────
thin = Border(
    left=Side(style='thin', color='D0D5DD'),
    right=Side(style='thin', color='D0D5DD'),
    top=Side(style='thin', color='D0D5DD'),
    bottom=Side(style='thin', color='D0D5DD'),
)
fill_title = PatternFill('solid', fgColor='0F3D5E')
fill_section = PatternFill('solid', fgColor='1B6CA8')
fill_kpi = PatternFill('solid', fgColor='E8F1F8')
fill_kpi_warn = PatternFill('solid', fgColor='FFF4E0')
fill_kpi_good = PatternFill('solid', fgColor='E3F6EF')
fill_kpi_hot = PatternFill('solid', fgColor='FDE8E6')
fill_header = PatternFill('solid', fgColor='1B6CA8')
fill_alt = PatternFill('solid', fgColor='F7F9FB')
fill_insight = PatternFill('solid', fgColor='F0F7FC')
fill_note = PatternFill('solid', fgColor='FFF8EB')
fill_edit = PatternFill('solid', fgColor='FFF9E8')
fill_edit_hdr = PatternFill('solid', fgColor='E07A3D')
fill_panel = PatternFill('solid', fgColor='F5F7FA')
font_title = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
font_section = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
font_kpi_label = Font(name='Calibri', size=9, bold=True, color='0F3D5E')
font_kpi_value = Font(name='Calibri', size=16, bold=True, color='0F3D5E')
font_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
font_body = Font(name='Calibri', size=10, color='1C2833')
font_insight = Font(name='Calibri', size=10, italic=True, color='1C2833')
font_note = Font(name='Calibri', size=9, color='7D6608')
font_edit = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
font_hint = Font(name='Calibri', size=9, italic=True, color='667085')


def style_header_row(ws_, row, start_col, end_col, fill=None):
    for col in range(start_col, end_col + 1):
        cell = ws_.cell(row=row, column=col)
        cell.fill = fill or fill_header
        cell.font = font_header
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = thin


def write_table(ws_, start_row, start_col, df):
    headers = list(df.columns)
    for j, h in enumerate(headers):
        ws_.cell(row=start_row, column=start_col + j, value=h)
    style_header_row(ws_, start_row, start_col, start_col + len(headers) - 1)
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            cell = ws_.cell(row=start_row + i, column=start_col + j, value=val if pd.notna(val) else None)
            cell.font = font_body
            cell.border = thin
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if i % 2 == 0:
                cell.fill = fill_alt
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
            elif isinstance(val, (int, np.integer)):
                cell.number_format = '#,##0'
    return start_row + len(df)


def colorize_series(chart, colors=None):
    colors = colors or PALETTE
    if not chart.series:
        return
    series = chart.series[0]
    pts = []
    for i, color in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties = GraphicalProperties(solidFill=color)
        pts.append(pt)
    series.data_points = pts


def apply_bar_layout(chart, width=16, height=9):
    """Modern bar chart: no legend clutter, outside labels, roomy plot area."""
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.legend = None
    chart.gapWidth = 80
    chart.overlap = 0
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showPercent = False
    # Leave space for title + category labels
    chart.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.18, y=0.12, w=0.78, h=0.78,
        )
    )
    if chart.x_axis is not None:
        chart.x_axis.delete = False
        chart.x_axis.majorTickMark = 'out'
        chart.x_axis.minorTickMark = 'none'
    if chart.y_axis is not None:
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines = None  # set below carefully
    return chart


def apply_doughnut_layout(chart, width=14, height=9):
    """Doughnut with bottom legend — avoids slice label overlap."""
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.holeSize = 58
    chart.legend = Legend(legendPos='b', overlay=False)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False  # categories in legend → no overlap on slices
    chart.dataLabels.showSerName = False
    chart.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.12, y=0.08, w=0.76, h=0.62,
        )
    )
    return chart


def write_edit_block(ws_, row, col, title, headers, rows, hint='EDIT YELLOW CELLS -> chart updates'):
    """Write a clearly labeled editable data block. Returns (header_row, last_data_row, n_cols)."""
    ws_.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(headers) - 1)
    cell = ws_.cell(row=row, column=col, value=title)
    cell.font = font_edit
    cell.fill = fill_edit_hdr
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(col, col + len(headers)):
        ws_.cell(row=row, column=c).fill = fill_edit_hdr
        ws_.cell(row=row, column=c).border = thin

    ws_.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + len(headers) - 1)
    hint_cell = ws_.cell(row=row + 1, column=col, value=hint)
    hint_cell.font = font_hint
    hint_cell.fill = fill_edit

    hdr_row = row + 2
    for j, h in enumerate(headers):
        cell = ws_.cell(row=hdr_row, column=col + j, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')

    for i, rdata in enumerate(rows):
        for j, val in enumerate(rdata):
            cell = ws_.cell(row=hdr_row + 1 + i, column=col + j, value=val)
            cell.font = font_body
            cell.border = thin
            cell.fill = fill_edit
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
            elif isinstance(val, int):
                cell.number_format = '#,##0'
    last = hdr_row + len(rows)
    return hdr_row, last, len(headers)


# ── Build workbook ────────────────────────────────────────────────
wb = load_workbook(path)
for name in ['Executive_Dashboard', 'Chart_Data', 'Charts']:
    if name in wb.sheetnames:
        del wb[name]

ws = wb.create_sheet('Executive_Dashboard', 0)
ch = wb.create_sheet('Charts', 1)

# ═══════════════════════ Executive_Dashboard ══════════════════════
widths = {
    'A': 3, 'B': 28, 'C': 22, 'D': 18, 'E': 36, 'F': 36, 'G': 22, 'H': 14,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.merge_cells('B2:G2')
ws['B2'] = 'Duplicate Index Analysis — Executive Dashboard'
ws['B2'].font = font_title
ws['B2'].fill = fill_title
ws['B2'].alignment = Alignment(vertical='center', horizontal='left')
for col in range(2, 8):
    ws.cell(row=2, column=col).fill = fill_title
ws.row_dimensions[2].height = 32

ws.merge_cells('B3:G3')
ws['B3'] = (
    f'Generated {datetime.now().strftime("%Y-%m-%d %H:%M")}  |  '
    f'Scope: {kpi_dbs_scanned} databases, {kpi_indexes:,} indexes  |  '
    f'Charts (editable) → see "Charts" sheet'
)
ws['B3'].font = Font(name='Calibri', size=9, color='566573')

ws.merge_cells('B5:G5')
ws['B5'] = 'KEY PERFORMANCE INDICATORS'
ws['B5'].font = font_section
ws['B5'].fill = fill_section
for col in range(2, 8):
    ws.cell(row=5, column=col).fill = fill_section

kpis = [
    (6, 2, 'Databases Scanned', kpi_dbs_scanned, fill_kpi),
    (6, 3, 'Indexes Scanned', kpi_indexes, fill_kpi),
    (6, 4, 'Duplicate Pairs', kpi_pairs, fill_kpi_hot),
    (6, 5, 'DBs with Duplicates', kpi_dbs_with_dups, fill_kpi_warn),
    (6, 6, 'Dup Reclaim (MB)', round(kpi_reclaim_dup, 2), fill_kpi_good),
    (6, 7, 'Safe Drop Candidates', kpi_safe, fill_kpi_good),
    (8, 2, 'Exact Duplicates', kpi_exact, fill_kpi_hot),
    (8, 3, 'Duplicate Key', kpi_dupkey, fill_kpi_warn),
    (8, 4, 'Left-Prefix Redundant', kpi_prefix, fill_kpi),
    (8, 5, 'Need Usage Review', kpi_review, fill_kpi_warn),
    (8, 6, 'Consol. Suggestions', kpi_consol, fill_kpi),
    (8, 7, 'Consol. Reclaim (MB)', round(kpi_consol_reclaim, 2), fill_kpi_good),
]
for r, c, label, value, fill in kpis:
    ws.cell(row=r, column=c, value=label).font = font_kpi_label
    ws.cell(row=r, column=c).fill = fill
    ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=c).border = thin
    cell = ws.cell(row=r + 1, column=c, value=value)
    cell.font = font_kpi_value
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin
    if isinstance(value, float):
        cell.number_format = '#,##0.00'
    elif isinstance(value, int) and value > 999:
        cell.number_format = '#,##0'
    ws.row_dimensions[r + 1].height = 28

ws.merge_cells('B10:G10')
ws['B10'] = "BIRD'S-EYE FINDINGS"
ws['B10'].font = font_section
ws['B10'].fill = fill_section
for col in range(2, 8):
    ws.cell(row=10, column=col).fill = fill_section

insights = [
    f'1. SCALE: {kpi_pairs:,} duplicate-index pairs across {kpi_dbs_with_dups} of {kpi_dbs_scanned} databases '
    f'({100 * kpi_dbs_with_dups / kpi_dbs_scanned:.0f}% of estate). Estimated reclaim ≈ {kpi_reclaim_dup:.1f} MB (duplicates only).',
    f'2. TYPE MIX: Left Prefix dominates ({kpi_prefix} / {100 * kpi_prefix / kpi_pairs:.0f}%). '
    f'Exact Dup ({kpi_exact}) = highest-confidence drops; Dup Key ({kpi_dupkey}) needs leading-key validation.',
    f'3. SAFETY: {kpi_safe} pairs ({100 * kpi_safe / kpi_pairs:.0f}%) unused/write-only → Wave 1 candidates. '
    f'{kpi_review} still have reads → validate with Query Store.',
    '4. SCHEMA DEBT: A few patterns hit 100–200 DBs (AccBillDetail exact, SalesReturn prefix, MealTracking key-order, '
    'LabTestReport prefix). Fix once in product schema → multiply impact.',
    f'5. CONSOLIDATION: {kpi_consol} big-table suggestions ≈ {kpi_consol_reclaim:.0f} MB (mostly Leading Key review). '
    f'Higher risk/payoff — treat separately from exact duplicates.',
    f'6. COMBINED OPPORTUNITY: Duplicates (~{kpi_reclaim_dup:.0f} MB) + consolidation (~{kpi_consol_reclaim:.0f} MB) '
    f'≈ {kpi_reclaim_dup + kpi_consol_reclaim:.0f} MB. Bigger win: fewer indexes, less write amp, cleaner plans.',
    '7. CHARTS: Open the "Charts" sheet - each chart has a yellow editable data table beside it. Change a value -> chart updates.',
]
for i, text in enumerate(insights):
    row = 11 + i
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = font_insight
    cell.fill = fill_insight
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 34
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = fill_insight
        ws.cell(row=row, column=col).border = thin

# Detail tables (start earlier now that charts moved off this sheet)
row = 20
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
ws.cell(row=row, column=2, value='TABLE — Duplicate Pairs & Reclaim by Type').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = fill_section
by_type_out = by_type[['duplicate_type', 'pair_count', 'reclaim_mb']].rename(columns={
    'duplicate_type': 'Duplicate Type', 'pair_count': 'Pair Count', 'reclaim_mb': 'Reclaimable MB'
})
by_type_out['% of Pairs'] = (by_type_out['Pair Count'] / by_type_out['Pair Count'].sum() * 100).round(1)
write_table(ws, row + 1, 2, by_type_out)

row = 26
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
ws.cell(row=row, column=2, value='TABLE — Safety / Usage Bucket').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, by_safety.rename(columns={
    'safety': 'Safety Bucket', 'pairs': 'Pairs', 'reclaim_mb': 'Reclaimable MB'
}))

row = 32
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
ws.cell(row=row, column=2, value='TABLE — Environment Breakdown').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 7):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, by_env.rename(columns={
    'env': 'Environment', 'pairs': 'Pairs', 'databases': 'Databases', 'reclaim_mb': 'Reclaimable MB'
}))

row = 44
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
ws.cell(row=row, column=2, value='TABLE — Top Databases by Reclaimable MB (Duplicates)').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, top_dbs[['database_name', 'pairs', 'reclaim_mb']].rename(columns={
    'database_name': 'Database', 'pairs': 'Pairs', 'reclaim_mb': 'Reclaimable MB'
}))

row = 60
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
ws.cell(row=row, column=2, value='TABLE — Cross-DB Schema Debt Patterns').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 8):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, patterns[
    ['table_name', 'duplicate_type', 'index_keep', 'index_drop_candidate', 'databases_affected', 'reclaim_mb']
].rename(columns={
    'table_name': 'Table', 'duplicate_type': 'Type', 'index_keep': 'Keep Index',
    'index_drop_candidate': 'Drop Candidate', 'databases_affected': 'DBs Affected', 'reclaim_mb': 'Reclaim MB'
}))

row = 72
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
ws.cell(row=row, column=2, value='TABLE — Priority Action Queue (Top 20)').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 9):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, top_actions.rename(columns={
    'database_name': 'Database', 'table_name': 'Table', 'duplicate_type': 'Type',
    'index_drop_candidate': 'Drop Index', 'index_keep': 'Keep Index',
    'drop_size_f': 'Size MB', 'drop_usage_f': 'Drop Usage', 'safety': 'Safety',
    'action_f': 'Suggested Action', 'priority_score': 'Score'
})[['Database', 'Table', 'Type', 'Drop Index', 'Keep Index', 'Size MB', 'Safety', 'Score']])

row = 96
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
ws.cell(row=row, column=2, value='TABLE — Consolidation by Type').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 7):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, consol_by_type[['suggestion_type', 'suggestions', 'actionable', 'reclaim_mb']].rename(columns={
    'suggestion_type': 'Suggestion Type', 'suggestions': 'Suggestions',
    'actionable': 'Actionable', 'reclaim_mb': 'Est. Reclaim MB'
}))

row = 103
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
ws.cell(row=row, column=2, value='TABLE — Top DBs by Consolidation Reclaim').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, top_consol_dbs.rename(columns={
    'database_name': 'Database', 'suggestions': 'Suggestions', 'reclaim_mb': 'Est. Reclaim MB'
}))

row = 117
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
ws.cell(row=row, column=2, value='TABLE — Largest Tables in Scope').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 9):
    ws.cell(row=row, column=col).fill = fill_section
write_table(ws, row + 1, 2, big_top.rename(columns={
    'database_name': 'Database', 'schema_name': 'Schema', 'table_name': 'Table',
    'table_rows': 'Rows', 'table_size_mb': 'Size MB', 'index_count': 'Indexes', 'nc_index_count': 'NC Indexes'
}))

row = 133
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
ws.cell(row=row, column=2, value='RECOMMENDED REMEDIATION PLAYBOOK').font = font_section
ws.cell(row=row, column=2).fill = fill_section
for col in range(2, 8):
    ws.cell(row=row, column=col).fill = fill_section
playbook = [
    'Wave 1 (Low risk): Exact Dup + Safe. Start with AccBillDetail IX_AccBillDetail_BillId (~203 DBs). Validate non-prod, then roll.',
    'Wave 2 (Medium): Left Prefix where drop is unused (SalesReturn, AccBill BillType, AccBillDetail MasterId, LabTestReport labId).',
    'Wave 3 (Careful): Dup Key / key-order (MealTracking UQ, AccBill Stock_Filter vs BillType). Use Query Store for seeks.',
    'Wave 4: Consolidation — Prefix Subsumed & Drop Redundant first; Leading Key = design review, not blind drop.',
    'Structural fix: Promote Wave 1–2 into base product schema so new tenants stop inheriting the debt.',
    'Ops: Usage stats reset on restart. If uptime is short, re-check after a full business cycle / use Query Store.',
]
for i, text in enumerate(playbook):
    r = row + 1 + i
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = font_body
    cell.fill = fill_note
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 36
    for col in range(2, 8):
        ws.cell(row=r, column=col).fill = fill_note
        ws.cell(row=r, column=col).border = thin

ws.freeze_panes = 'B5'

# ═══════════════════════ Charts sheet ═════════════════════════════
# Layout: each panel = [editable yellow table on left] + [chart on right]
# Vertical spacing ~22 rows per panel to prevent chart overlap.

for col, w in {
    'A': 3, 'B': 22, 'C': 12, 'D': 12, 'E': 3, 'F': 18, 'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 14,
}.items():
    ch.column_dimensions[col].width = w

ch.merge_cells('B2:K2')
ch['B2'] = 'Interactive Charts — edit yellow cells; charts update automatically'
ch['B2'].font = font_title
ch['B2'].fill = fill_title
for col in range(2, 12):
    ch.cell(row=2, column=col).fill = fill_title
ch.row_dimensions[2].height = 30

ch.merge_cells('B3:K3')
ch['B3'] = (
    'How to edit: change Category or Value in the yellow tables -> chart refreshes. '
    'To add a row: insert inside the yellow block, then right-click chart -> Select Data -> resize the range. '
    'To restyle: click chart -> Chart Design ribbon.'
)
ch['B3'].font = font_hint
ch.row_dimensions[3].height = 28

# ── Panel 1: Pairs by Type (doughnut) + Reclaim by Type (column) ──
panel_row = 5
ch.merge_cells(start_row=panel_row, start_column=2, end_row=panel_row, end_column=11)
ch.cell(row=panel_row, column=2, value='1 - Duplicate Type Mix').font = font_section
ch.cell(row=panel_row, column=2).fill = fill_section
for c in range(2, 12):
    ch.cell(row=panel_row, column=c).fill = fill_section

type_rows = [
    (r.label, int(r.pair_count), round(float(r.reclaim_mb), 2))
    for r in by_type.itertuples(index=False)
]
hdr1, end1, _ = write_edit_block(
    ch, panel_row + 1, 2,
    'DATA - Pairs & Reclaim by Type',
    ['Category', 'Pairs', 'Reclaim MB'],
    type_rows,
)

donut1 = DoughnutChart()
donut1.title = 'Pairs by Type'
donut1.add_data(Reference(ch, min_col=3, min_row=hdr1, max_row=end1), titles_from_data=True)
donut1.set_categories(Reference(ch, min_col=2, min_row=hdr1 + 1, max_row=end1))
apply_doughnut_layout(donut1, width=13, height=9)
colorize_series(donut1)
ch.add_chart(donut1, 'F7')

bar1 = BarChart()
bar1.type = 'col'
bar1.title = 'Reclaim MB by Type'
bar1.y_axis.title = None  # avoid axis-title overlap; unit is in chart title
bar1.add_data(Reference(ch, min_col=4, min_row=hdr1, max_row=end1), titles_from_data=True)
bar1.set_categories(Reference(ch, min_col=2, min_row=hdr1 + 1, max_row=end1))
apply_bar_layout(bar1, width=13, height=9)
# column chart needs different plot margins (categories at bottom)
bar1.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.08, y=0.14, w=0.86, h=0.72))
colorize_series(bar1)
ch.add_chart(bar1, 'I7')

# ── Panel 2: Safety doughnut + Env bar ────────────────────────────
panel_row = 28
ch.merge_cells(start_row=panel_row, start_column=2, end_row=panel_row, end_column=11)
ch.cell(row=panel_row, column=2, value='2 - Safety & Environment').font = font_section
ch.cell(row=panel_row, column=2).fill = fill_section
for c in range(2, 12):
    ch.cell(row=panel_row, column=c).fill = fill_section

safety_rows = [(r.safety, int(r.pairs), round(float(r.reclaim_mb), 2)) for r in by_safety.itertuples(index=False)]
hdr2, end2, _ = write_edit_block(
    ch, panel_row + 1, 2,
    'DATA - Usage Safety',
    ['Category', 'Pairs', 'Reclaim MB'],
    safety_rows,
)

env_rows = [(r.env, int(r.pairs), round(float(r.reclaim_mb), 2)) for r in by_env.itertuples(index=False)]
hdr2b, end2b, _ = write_edit_block(
    ch, panel_row + 1, 6,
    'DATA - Environment',
    ['Category', 'Pairs', 'Reclaim MB'],
    env_rows,
)

donut2 = DoughnutChart()
donut2.title = 'Drop Candidates by Safety'
donut2.add_data(Reference(ch, min_col=3, min_row=hdr2, max_row=end2), titles_from_data=True)
donut2.set_categories(Reference(ch, min_col=2, min_row=hdr2 + 1, max_row=end2))
apply_doughnut_layout(donut2, width=13, height=9)
colorize_series(donut2, ['2A9D8F', 'E07A3D', '8A8A8A'])
ch.add_chart(donut2, 'B40')

bar2 = BarChart()
bar2.type = 'bar'
bar2.title = 'Pairs by Environment'
bar2.add_data(Reference(ch, min_col=7, min_row=hdr2b, max_row=end2b), titles_from_data=True)
bar2.set_categories(Reference(ch, min_col=6, min_row=hdr2b + 1, max_row=end2b))
apply_bar_layout(bar2, width=15, height=9)
colorize_series(bar2, ['1B6CA8'] * 10)
ch.add_chart(bar2, 'F40')

# ── Panel 3: Top DBs + Schema patterns ────────────────────────────
panel_row = 60
ch.merge_cells(start_row=panel_row, start_column=2, end_row=panel_row, end_column=11)
ch.cell(row=panel_row, column=2, value='3 - Hotspots & Schema Debt').font = font_section
ch.cell(row=panel_row, column=2).fill = fill_section
for c in range(2, 12):
    ch.cell(row=panel_row, column=c).fill = fill_section

db_rows = [(r.label, int(r.pairs), round(float(r.reclaim_mb), 2)) for r in top_dbs.itertuples(index=False)]
hdr3, end3, _ = write_edit_block(
    ch, panel_row + 1, 2,
    'DATA - Top DBs by Reclaim MB',
    ['Database', 'Pairs', 'Reclaim MB'],
    db_rows,
)

pat_rows = [
    (r.label, int(r.databases_affected), round(float(r.reclaim_mb), 2))
    for r in patterns.itertuples(index=False)
]
hdr3b, end3b, _ = write_edit_block(
    ch, panel_row + 1, 6,
    'DATA - Cross-DB Patterns',
    ['Pattern', 'DBs Hit', 'Reclaim MB'],
    pat_rows,
)

bar3 = BarChart()
bar3.type = 'bar'
bar3.title = 'Top Databases — Reclaim MB'
bar3.add_data(Reference(ch, min_col=4, min_row=hdr3, max_row=end3), titles_from_data=True)
bar3.set_categories(Reference(ch, min_col=2, min_row=hdr3 + 1, max_row=end3))
apply_bar_layout(bar3, width=16, height=11)
# more left margin for DB names
bar3.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.28, y=0.10, w=0.68, h=0.82))
colorize_series(bar3, ['E07A3D'] * 15)
ch.add_chart(bar3, 'B78')

bar4 = BarChart()
bar4.type = 'bar'
bar4.title = 'Schema Patterns — DBs Affected'
bar4.add_data(Reference(ch, min_col=7, min_row=hdr3b, max_row=end3b), titles_from_data=True)
bar4.set_categories(Reference(ch, min_col=6, min_row=hdr3b + 1, max_row=end3b))
apply_bar_layout(bar4, width=16, height=11)
bar4.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.32, y=0.10, w=0.64, h=0.82))
colorize_series(bar4, ['2A9D8F'] * 10)
ch.add_chart(bar4, 'F78')

# ── Panel 4: Consolidation + Opportunity ──────────────────────────
panel_row = 102
ch.merge_cells(start_row=panel_row, start_column=2, end_row=panel_row, end_column=11)
ch.cell(row=panel_row, column=2, value='4 - Consolidation Opportunity').font = font_section
ch.cell(row=panel_row, column=2).fill = fill_section
for c in range(2, 12):
    ch.cell(row=panel_row, column=c).fill = fill_section

consol_rows = [
    (r.label, int(r.suggestions), round(float(r.reclaim_mb), 2))
    for r in consol_by_type.itertuples(index=False)
]
hdr4, end4, _ = write_edit_block(
    ch, panel_row + 1, 2,
    'DATA - Consolidation by Type',
    ['Category', 'Suggestions', 'Reclaim MB'],
    consol_rows,
)

opp_rows = [
    ('Dup drops', round(kpi_reclaim_dup, 2)),
    ('Consolidation', round(kpi_consol_reclaim, 2)),
]
hdr4b, end4b, _ = write_edit_block(
    ch, panel_row + 1, 6,
    'DATA - Total Opportunity (MB)',
    ['Category', 'Reclaim MB'],
    opp_rows,
)

bar5 = BarChart()
bar5.type = 'col'
bar5.title = 'Consolidation Reclaim by Type'
bar5.add_data(Reference(ch, min_col=4, min_row=hdr4, max_row=end4), titles_from_data=True)
bar5.set_categories(Reference(ch, min_col=2, min_row=hdr4 + 1, max_row=end4))
apply_bar_layout(bar5, width=13, height=9)
bar5.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.08, y=0.14, w=0.86, h=0.72))
colorize_series(bar5)
ch.add_chart(bar5, 'B116')

bar6 = BarChart()
bar6.type = 'col'
bar6.title = 'Duplicates vs Consolidation (MB)'
bar6.add_data(Reference(ch, min_col=7, min_row=hdr4b, max_row=end4b), titles_from_data=True)
bar6.set_categories(Reference(ch, min_col=6, min_row=hdr4b + 1, max_row=end4b))
apply_bar_layout(bar6, width=13, height=9)
bar6.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.08, y=0.14, w=0.86, h=0.72))
colorize_series(bar6, PALETTE_ALT)
ch.add_chart(bar6, 'F116')

# Tips footer
tip_row = 136
ch.merge_cells(start_row=tip_row, start_column=2, end_row=tip_row, end_column=11)
ch.cell(row=tip_row, column=2, value='EDITING TIPS').font = font_section
ch.cell(row=tip_row, column=2).fill = fill_section
for c in range(2, 12):
    ch.cell(row=tip_row, column=c).fill = fill_section

tips = [
    '- Yellow cells are the chart source - edit Category text or numbers; the chart reflects new values.',
    '- Short category names prevent label overlap. Keep labels under ~28 characters for horizontal bars.',
    '- Doughnuts show % on slices and full names in the bottom legend (no overlapping slice text).',
    '- Bar charts hide the legend (categories are on the axis) and place values at the end of each bar.',
    '- To change chart type/colors: click the chart -> Chart Design -> Change Chart Type / Change Colors.',
    '- To expand a series after adding rows: right-click chart -> Select Data -> edit the value/category range.',
]
for i, t in enumerate(tips):
    r = tip_row + 1 + i
    ch.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    cell = ch.cell(row=r, column=2, value=t)
    cell.font = font_note
    cell.fill = fill_note
    for c in range(2, 12):
        ch.cell(row=r, column=c).fill = fill_note

ch.freeze_panes = 'B5'
ch.sheet_view.showGridLines = False
ws.sheet_view.showGridLines = False

wb.save(path)
print('Saved OK')
print('Sheets:', wb.sheetnames)
print(f'Charts sheet panels ready; reclaim_dup={kpi_reclaim_dup}, consol={kpi_consol_reclaim}')
